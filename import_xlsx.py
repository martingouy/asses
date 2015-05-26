from openpyxl import load_workbook
import os
import json
import traceback
import sys

def importation(file):
    
    try:
        wb = load_workbook(filename=file, read_only=True)
        
    
        mySession={'attributes':[], 'k_calculus':[{'method':'multiplicative', 'active':True, 'k':[], 'GK':None},{'method':'multilinear','active':False, 'k':[], 'GK':None}]}
        
        
        for sheet in wb:
            myAttribut={}
            ws = wb[sheet.title] # ws is now an IterableWorksheet
            if sheet.title=="Multi attribute multilinear" or sheet.title=="Multi attribute multiplicative":
                continue
            
            myAttribut['name']=ws['B2'].value
            myAttribut['unit']=ws['B3'].value
            myAttribut['val_min']=ws['B4'].value
            myAttribut['val_max']=ws['B5'].value
            myAttribut['method']=ws['B6'].value
            myAttribut['mode']=ws['B7'].value
            myAttribut['checked']=ws['B8'].value
            
            myAttribut['questionnaire']={};
            
            
                                                              
            ligne=3
            number=0
            mesPoints=[]
            
            while ws['C'+str(ligne)].value!=None:
                mesPoints.append([ws['C'+str(ligne)].value, ws['D'+str(ligne)].value])
                ligne=ligne+1
                number=number+1
            
            myAttribut['questionnaire']['points']=mesPoints;
            myAttribut['questionnaire']['number']=number;
            
            mySession['attributes'].append(myAttribut)


        for sheet in wb:
            ws = wb[sheet.title] # ws is now an IterableWorksheet
            if sheet.title!="Multi attribute multilinear" and sheet.title!="Multi attribute multiplicative":
                continue
            
            ligne=2
            mesK=[]
            while ws['C'+str(ligne)].value!=None:
                mesK.append({'ID':ws['A'+str(ligne)].value,'ID_attribute':json.loads(ws['D'+str(ligne)].value),'attribute':json.loads(ws['C'+str(ligne)].value), 'value':ws['B'+str(ligne)].value})
                ligne=ligne+1
        
            
            if sheet.title=="Multi attribute multilinear":
                mySession['k_calculus'][1]['k']=mesK
                mySession['k_calculus'][1]['GK']=ws['B'+str(ligne)].value
            
            elif sheet.title=="Multi attribute multiplicative":
                mySession['k_calculus'][0]['k']=mesK
                mySession['k_calculus'][0]['GK']=ws['B'+str(ligne)].value
        
     
        os.remove(file)
        return {'success':True, 'data':mySession}
    except Exception, err:
        os.remove(file)
        return {'success':False, 'data':traceback.format_exc()}





